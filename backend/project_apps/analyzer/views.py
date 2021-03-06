from django.shortcuts import render
# 从models中引入
from project_apps.analyzer.models import Analyzer
# 返回Json格式的数据
from django.http import JsonResponse

import json
import os
from django.conf import settings
import openpyxl

import uuid
import hashlib
# 引入Q 查询
from django.db.models import Q


# Create your views here.
# 使用ORM，获取所有学生的接口，
def get_student(request):
    try:
        analyzer_object = Analyzer.objects.all().values()
        analyzer = list(analyzer_object)
        return JsonResponse({'code': 1, 'data': analyzer})
    except Exception as e:
        return JsonResponse({'code': 0, 'msg': '获取学生信息异常' + str(e)})


# 使用ORM查询学生信息
def query_student(request):
    # axios 默认数据为json格式--使用inputstr存储
    data = json.loads(request.body.decode("utf-8"))
    try:
        # 获取满足条件的学生信息，并转为字典格式
        analyzer_object = Analyzer.objects.filter(
            Q(email__icontains=data['inputstr']) | Q(name__icontains=data['inputstr']) |
            Q(gender__icontains=data['inputstr']) | Q(address__icontains=data['inputstr']) |
            Q(mobile__icontains=data['inputstr']) | Q(address__icontains=data['inputstr'])).values()
        analyzer = list(analyzer_object)
        return JsonResponse({'code': 1, 'data': analyzer})
    except Exception as e:
        return JsonResponse({'code': 0, 'msg': '获取学生信息异常' + str(e)})


# 检查学生学号是否已存在
def check_student(request):
    data = json.loads(request.body.decode("utf-8"))
    try:
        students_object = Analyzer.objects.filter(sno=data['sno'])
        if students_object.count() == 0:
            return JsonResponse({'code': 1, 'exists': False})
        else:
            return JsonResponse({'code': 1, 'exists': True})
    except Exception as e:
        return JsonResponse({'code': 0, 'msg': "检查学生信息出错" + str(e)})


# 添加学生
def add_student(request):
    data = json.loads(request.body.decode('utf-8'))
    try:
        #     添加学生信息
        student_obj = Analyzer(sno=data['sno'], name=data['name'], birthday=data['birthday'], gender=data['gender'],
                               mobile=data['mobile'], email=data['email'], address=data['address'], image=data['image'])
        student_obj.save()
        analyzer_object = Analyzer.objects.all().values()
        analyzer = list(analyzer_object)
        return JsonResponse({'code': 1, 'data': analyzer})
    except Exception as e:
        return JsonResponse({'code': 0, 'msg': '添加学生失败，具体原因' + str(e)})


# 修改学生
def update_student(request):
    data = json.loads(request.body.decode('utf-8'))
    try:
        #     添加学生信息
        student_obj = Analyzer.objects.get(sno=data['sno'])
        student_obj.name = data['name']
        student_obj.gender = data['gender']
        student_obj.birthday = data['birthday']
        student_obj.mobile = data['mobile']
        student_obj.email = data['email']
        student_obj.address = data['address']
        student_obj.save()
        analyzer_object = Analyzer.objects.all().values()
        analyzer = list(analyzer_object)
        return JsonResponse({'code': 1, 'data': analyzer})
    except Exception as e:
        return JsonResponse({'code': 0, 'msg': '修改学生失败，具体原因' + str(e)})


# 删除学生信息
def delete_student(request):
    data = json.loads(request.body.decode('utf-8'))
    try:
        # 获取学生信息
        student_obj = Analyzer.objects.get(sno=data['sno'])
        # 删除操作
        student_obj.delete()
        analyzer_object = Analyzer.objects.all().values()
        analyzer = list(analyzer_object)
        return JsonResponse({'code': 1, 'data': analyzer})
    except Exception as e:
        return JsonResponse({'code': 0, 'msg': '删除学生失败，具体原因' + str(e)})


# 批量删除学生信息
def delete_students(request):
    data = json.loads(request.body.decode('utf-8'))
    try:
        # 获取学生信息
        for one_student in data['student']:
            obj_student = Analyzer.objects.get(sno=one_student['sno'])
            obj_student.delete()
        # 删除后获取所有学生信息
        analyzer_object = Analyzer.objects.all().values()
        analyzer = list(analyzer_object)
        return JsonResponse({'code': 1, 'data': analyzer})
    except Exception as e:
        return JsonResponse({'code': 0, 'msg': '批量删除学生失败，具体原因' + str(e)})


# 上传头像
def avatar_upload(request):
    # 接收上传的文件
    rev_file = request.FILES.get('avatar')
    # 判断是否有文件
    if not rev_file:
        return JsonResponse({'code': 0, 'msg': '上传的图片不存在！'})
    # 获得唯一名字
    new_name = get_random_str()
    # 写入
    file_path = os.path.join(settings.MEDIA_ROOT, new_name + os.path.splitext(rev_file.name)[1])
    try:
        f = open(file_path, 'wb')
        for i in rev_file.chunks():
            f.write(i)
        f.close()
        return JsonResponse({'code': 1, 'name': new_name + os.path.splitext(rev_file.name)[1]})
    except Exception as e:
        return JsonResponse({'code': 0, 'msg': str(e)})


def get_random_str():
    # 获取uuid 随机数
    uuid_val = uuid.uuid4()
    # 获取uuid随机数字符串
    uuid_str = str(uuid_val).encode('utf-8')
    # 获取md5实例
    md5 = hashlib.md5()
    # 拿取uuid的md5摘要
    md5.update(uuid_str)
    # 返回固定长度的字符串
    return md5.hexdigest()


# excel的导入
def excel_import(request):
    # 1.获得上传的文件
    rev_file = request.FILES.get('excel')
    # 判断是否有文件
    if not rev_file:
        return JsonResponse({'code': 0, 'msg': '上传的excel文件不存在！'})
    # 获得唯一名字
    new_name = get_random_str()
    # 写入
    file_path = os.path.join(settings.MEDIA_ROOT, new_name + os.path.splitext(rev_file.name)[1])
    try:
        f = open(file_path, 'wb')
        for i in rev_file.chunks():
            f.write(i)
        f.close()

    except Exception as e:
        return JsonResponse({'code': 0, 'msg': str(e)})
    #2. 打开文件获得数据
    ex_students = get_excel_data(file_path)
    # 准备数据
    success = 0
    error = 0
    errors = []
    #3. 将获得的数据写入数据库
    for one_student in ex_students:
        try:
            obj = Analyzer.objects.create(sno=one_student['sno'], name=one_student['name'], gender=one_student['gender'],
                                                      birthday=one_student['birthday'],mobile=one_student['mobile'],
                                                   email=one_student['email'],address=one_student['address'])
            obj.save()
            success += 1
        except:
            error +=1
            errors.append(one_student['sno'])

    #4. 获得所有学生的信息
    stu_obj = Analyzer.objects.all().values()
    stu_obj = list(stu_obj)
    return JsonResponse({'code':1, 'success':success,'error':error,'errors':error, 'data':stu_obj})

# excel 的导出
def excel_export(request):
    # 获取所有的学生信息
    obj_students = Analyzer.objects.all().values()
    # 转为List
    students = list(obj_students)
    # 准备名称
    excel_name = get_random_str() + ".xlsx"
    # 准备写入的路劲
    path = os.path.join(settings.MEDIA_ROOT, excel_name)
    # 写入到Excel
    write_to_excel(students, path)
    # 返回
    return JsonResponse({'code': 1, 'name': excel_name})



def get_excel_data(path:str):
    # 获取工作簿
    workbook = openpyxl.load_workbook(path)
    # 获取sheet
    sheet = workbook['student']
    # 准备数据
    students = []
    # 准备key
    stu_key = ['sno','name','gender','birthday','mobile','email','address']
    # 获取数据
    for one_student in sheet.rows:
        temp = {}
        # 组合key和value
        for index,cell in enumerate(one_student):
            temp[stu_key[index]] = cell.value
        students.append(temp)
    return students

def write_to_excel(data:list, path:str):
    """把数据库写入到Excel"""
    # 实例化一个workbook
    workbook = openpyxl.Workbook()
    # 激活一个sheet
    sheet = workbook.active
    # 为sheet命名
    sheet.title = 'student'
    # 准备keys
    keys = data[0].keys()
    # 准备写入数据
    for index, item in enumerate(data):
        # 遍历每一个元素
        for k, v in enumerate(keys):
            sheet.cell(row=index + 1, column=k + 1, value=str(item[v]))
    # 写入到文件
    workbook.save(path)